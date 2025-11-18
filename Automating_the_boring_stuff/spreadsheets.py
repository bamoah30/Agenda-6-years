# November 14, 2025
'''openpyxl module:
This module allows you to read and write Excel files (with .xlsx extension) using Python.
Some of the key functions and methods include:
.load_workbook() - opens an existing Excel file
.get_sheet_names() - lists all sheet names in the workbook
.get_sheet_by_name() - accesses a specific sheet by name
.workbook.active - gets the active sheet
.title - gets or sets the title of the sheet
.cell(row, column) - accesses a specific cell
NB: For the cell() method, the arguments are integers and both row and column indices start at 1, not 0.
.value - gets or sets the value of a cell
.save() - saves the workbook to a file
.max_row() - returns the highest row number that contains data
.max_column() - returns the highest column number that contains data

Example usage:
'''
import openpyxl
wb = openpyxl.Workbook().save('example.xlsx')  # Create a new workbook and save it
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet'] # Access sheet by name
print(sheet.title)  # Output >>> Sheet
cell = sheet.cell(row=1, column=2) # Access cell B1
print(cell.value)  # Output >>> 42 Depending on the content of cell B1
sheet.cell(row=2, column=2, value='Hello')  # Set cell
wb.save('example_modified.xlsx')  # Save the workbook with changes
print(sheet.max_row)  # Output >>> Maximum row number that contains data
print(sheet.max_column)  # Output >>> Maximum column number that contains data

# November 17, 2025
'''Spreadsheet Concepts Review:
1. Import the openpyxl module.
2. Call the openpyxl.load_workbook() function.
3. Get a Workbook object.
4. Call the get_active_sheet() or get_sheet_by_name() workbook method.
5. Get a Worksheet object.
6. Use indexing or the cell() sheet method with row and column keyword arguments.
7. Get a Cell object.
8. Read the Cell object’s value attribute.

Note:Whenever you edit a spreadsheet you’ve loaded from a file, you should 
always save the new, edited spreadsheet to a different filename than the 
original. That way, you’ll still have the original spreadsheet file to work with 
in case a bug in your code caused the new, saved file to have incorrect or 
corrupt data
Pravtical Examples  and Formulas:
'''
import openpyxl
wb = openpyxl.Workbook() # Create a new workbook
sheet = wb.active if wb.active else wb.create_sheet() # Get the active sheet, or create one if none exists
sheet['A1'] = 200 # Set cell A1 to 200
print(sheet['A1'].value)  # Output >>> 200

sheet['A2'] = 300 # Set cell A2 to 300
print(sheet['A2'].value)  # Output >>> 300

sheet['A3'] = '=SUM(A1:A2)' # Set cell A3 to the formula that sums A1 and A2
wb.save('writeFormula.xlsx')

# To read the value of a formula cell, you need to set data_only=True when loading the workbook or else you'll get the formula itself
wbFormulas = openpyxl.load_workbook('writeFormula.xlsx')
sheet = wbFormulas.active if wbFormulas.active else wbFormulas.worksheets[0] # Get the active sheet or first sheet
print(sheet['A3'].value) # Output >>> '=SUM(A1:A2)'

wbDataOnly = openpyxl.load_workbook('writeFormula.xlsx', data_only=True) # Load workbook with data_only=True
sheet = wbDataOnly.active if wbDataOnly.active else wbDataOnly.worksheets[0] # Get the active sheet

print(sheet['A3'].value) # Output >>> 500 # This output depends on  your chached value of the formula cell.


# November 18, 2025

'''Setting Row Heights and Column Widths:
Worksheet objects have row_dimensions and column_dimensions attributes that can control the height of rows and the width of columns.
NB1: The row_dimensions and column_dimensions attributes are dictionaries that map the row or column identifiers to Dimension objects.
NB2: Row heights are measured in points, where 1 point is 1/72 of an inch. The default row height is 15 points.
NB3: Column widths are measured in the number of characters that can fit in the cell using the default font. The default column width is 8.43 characters.

Example:'''
import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active if wb.active else wb.create_sheet() # Get the active sheet, or create one if none exists

sheet['A1'] = 'Tall row' # Set the value of cell A1 to 'Tall row'
sheet['B2'] = 'Wide column' # Set the value of cell B2 to 'Wide column'
sheet.row_dimensions[1].height = 70  # Set the height of row 1
sheet.column_dimensions['B'].width = 20  # Set the width of column B
wb.save('dimensions.xlsx')  # Save the workbook with the new dimensions

'''Merging and Unmerging Cells:
You can merge multiple cells into a single cell using the merge_cells() method and unmerge them using the unmerge_cells() method.
When merging cells, only the value of the top-left cell is retained; all other values are discarded.

Example: '''
import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active if wb.active else wb.create_sheet() # Get the active sheet, or create one if none exists
sheet.merge_cells('A1:D3')  # Merge cells from A1 to D3
sheet['A1'] = ' 12 Merged Cell'  # Set the value of the merged cell
sheet.merge_cells('C5:D5')  # Merge cells C5 and D5
sheet['C5'] = ' 2 Merged Cell'  # Set the value of the merged cell
wb.save('mergedCells.xlsx')  # Save the workbook with merged cells

#Unmerging cells
wb = openpyxl.load_workbook('mergedCells.xlsx')
sheet = wb.active if wb.active else wb.worksheets[0] # Get the active
sheet.unmerge_cells('A1:D3')  # Unmerge cells from A1 to D3
sheet.unmerge_cells('C5:D5')  # Unmerge cells C5 and D5
wb.save('unmergedCells.xlsx')  # Save the workbook with unmerged cells

'''Freezing Panes:
We freeze panes to keep specific rows or columns visible while scrolling through the rest of the worksheet.
You can freeze rows and columns in a worksheet using the freeze_panes attribute of the Worksheet object.
NB: Specifying a value to the freeze_panes attribute freezes all rows above and all columns to the left of the specified cell.
Example: '''
import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active if wb.active else wb.create_sheet() # Get the active sheet, or create one if none exists
sheet.freeze_panes = 'B2'  # Freeze the first row and first column
wb.save('frozenPanes.xlsx')  # Save the workbook with frozen panes

'''Charts:
OpenPyXL supports creating bar, line, scatter, and pie charts using the 
data in a sheet’s cells. 
To make a chart, you need to do the following:

1. Create a Reference object from a rectangular selection of cells.

2. Create a Series object by passing in the Reference object.

3. Create a Chart object.

4. Append the Series object to the Chart object.

5. Optionally, set the title, x and Y axis, etc.

6. Set the position of the Chart object.

7. Add the Chart object to the Worksheet object.
Reference objects are created by calling the openpyxl.charts.Reference() function and passing 
5 key  arguments:
1. The Worksheet object containing your chart data.
2. min_row - The first row of the data.
3. min_col - The first column of the data.
4. max_row - The last row of the data.
5. max_col - The last column of the data.

Example:'''
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import Series
wb = openpyxl.Workbook()
sheet = wb.active if wb.active else wb.create_sheet() # Get the active sheet, or create one if none exists

# Create some data in column A
for i in range(1, 11):
    sheet[f'A{i}'] = i

# Create a Reference object for the data in column A
refObj = Reference(sheet, min_row=1, min_col=1, max_row=10, max_col=1)  # From A1 to A10

# Create a Series object using the Reference object 
seriesObj = Series(refObj)

# Create a BarChart object
chartObj = BarChart()
chartObj.append(seriesObj)  # Append the Series to the Chart object
chartObj.title = 'My Bar Chart'
chartObj.x_axis.title = 'X Axis'
chartObj.y_axis.title = 'Y Axis'
chartObj.anchor = 'A12'  # Set the position of the chart to cell A12
sheet.add_chart(chartObj)  # Add the chart to the worksheet
